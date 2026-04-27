from __future__ import annotations

import datetime as dt
import re
import time
from collections import Counter, defaultdict

import openpyxl
import requests
from bs4 import BeautifulSoup


INPUT_WORKBOOK = r"C:\Users\student\Documents\golfModel\park_factors_2025_2026_ALL_FIXED.xlsx"
OUTPUT_WORKBOOK = r"C:\Users\student\Documents\golfModel\park_factors_2024_2026_ALL_FIXED.xlsx"

HEADERS = [
    "date",
    "stadium",
    "away_team",
    "home_team",
    "teams_playing",
    "hr_factor",
    "2b_3b_factor",
    "1b_factor",
    "runs_factor",
    "wind_receptive",
    "wind_hour1_mph",
    "wind_hour2_mph",
    "wind_hour3_mph",
    "air_hour1_temp",
    "air_hour2_temp",
    "air_hour3_temp",
    "humidity",
    "pressure",
]

UA = {
    "User-Agent": "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)"
}


def clean_value(value: str) -> str:
    return value.replace("\xa0", " ").strip()


def normalize_num(value: str) -> str:
    s = clean_value(value)
    if not s:
        return ""
    s = s.replace("%", "").replace("°", "").replace(",", "")
    if s in {"-", "—"}:
        return ""
    return s


def parse_game_blob(blob: str) -> tuple[str, str, str] | None:
    text = clean_value(blob)
    m = re.search(r"^(.*?)\s+\d{1,2}:\d{2}\s+([A-Z]{2,3})\s*@\s*([A-Z]{2,3})$", text)
    if not m:
        return None
    return clean_value(m.group(1)), clean_value(m.group(2)), clean_value(m.group(3))


def build_team_fallback(soup: BeautifulSoup) -> dict[str, tuple[str, str]]:
    park_to_teams: dict[str, set[str]] = defaultdict(set)
    tables = soup.find_all("table")
    if len(tables) < 2:
        return {}
    player_table = tables[1]
    rows = player_table.find_all("tr")
    for tr in rows[1:]:
        cells = [clean_value(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if len(cells) < 3:
            continue
        tm, park = cells[0], cells[2]
        if tm and park:
            park_to_teams[park].add(tm)
    out: dict[str, tuple[str, str]] = {}
    for park, teams in park_to_teams.items():
        if len(teams) >= 2:
            ordered = sorted(teams)
            out[park] = (ordered[0], ordered[1])
    return out


def parse_day(date_obj: dt.date) -> tuple[list[dict[str, str]], str | None]:
    date_str = date_obj.isoformat()
    url = f"https://www.ballparkpal.com/Park-Factors.php?date={date_str}"
    resp = requests.get(url, headers=UA, timeout=30)
    if resp.status_code != 200:
        return [], f"http_{resp.status_code}"

    soup = BeautifulSoup(resp.text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return [], "no_tables"

    game_table = tables[0]
    rows = game_table.find_all("tr")
    if len(rows) <= 1:
        return [], "no_game_rows"

    team_fallback = build_team_fallback(soup)
    parsed_rows: list[dict[str, str]] = []

    for tr in rows[1:]:
        cells = [clean_value(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if len(cells) < 17:
            continue
        if cells[0].lower() == "game":
            continue

        # Support both row formats:
        # 1) game blob in col 0 (percentages start col 1)
        # 2) percentages start in col 0 (no game blob)
        game_parts = parse_game_blob(cells[0])
        if game_parts is not None:
            stadium, away, home = game_parts
            offset = 1
        else:
            # If first column is a percentage, this row has no direct game blob.
            # We cannot safely map to a specific game row without stadium context.
            # Skip this row to avoid bad stadium/team assignments.
            continue

        if not (stadium and away and home):
            fb = team_fallback.get(stadium)
            if fb:
                away, home = fb

        if not (stadium and away and home):
            continue

        row = {
            "date": date_str,
            "stadium": stadium,
            "away_team": away,
            "home_team": home,
            "teams_playing": f"{away} @ {home}",
            "hr_factor": normalize_num(cells[offset + 0]),
            "2b_3b_factor": normalize_num(cells[offset + 1]),
            "1b_factor": normalize_num(cells[offset + 2]),
            "runs_factor": normalize_num(cells[offset + 3]),
            "wind_receptive": clean_value(cells[offset + 5]),
            "wind_hour1_mph": normalize_num(cells[offset + 7]),
            "wind_hour2_mph": normalize_num(cells[offset + 8]),
            "wind_hour3_mph": normalize_num(cells[offset + 9]),
            "air_hour1_temp": normalize_num(cells[offset + 10]),
            "air_hour2_temp": normalize_num(cells[offset + 11]),
            "air_hour3_temp": normalize_num(cells[offset + 12]),
            "humidity": normalize_num(cells[offset + 14]),
            "pressure": normalize_num(cells[offset + 15]),
        }
        parsed_rows.append(row)

    if not parsed_rows:
        return [], "no_parsed_rows"
    return parsed_rows, None


def load_base_rows(path: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(HEADERS, start=1):
            v = ws.cell(r, c).value
            rec[h] = "" if v is None else str(v).strip()
        rows.append(rec)
    return rows


def dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str, str]] = set()
    out: list[dict[str, str]] = []
    for row in rows:
        key = (
            row.get("date", ""),
            row.get("stadium", ""),
            row.get("away_team", ""),
            row.get("home_team", ""),
        )
        if key in seen:
            continue
        if not all(key):
            continue
        seen.add(key)
        out.append(row)
    return out


def write_workbook(path: str, rows: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "park_factors"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])
    wb.save(path)


def main() -> None:
    base_rows = load_base_rows(INPUT_WORKBOOK)
    base_count = len(base_rows)

    start = dt.date(2024, 3, 1)
    end = dt.date(2024, 11, 30)

    new_rows: list[dict[str, str]] = []
    included_dates: set[str] = set()
    skipped = Counter()

    d = start
    while d <= end:
        try:
            day_rows, reason = parse_day(d)
            if day_rows:
                new_rows.extend(day_rows)
                included_dates.add(d.isoformat())
            else:
                skipped[reason or "unknown"] += 1
        except Exception:
            skipped["exception"] += 1
        d += dt.timedelta(days=1)
        time.sleep(0.25)

    combined = dedupe_rows(base_rows + new_rows)
    write_workbook(OUTPUT_WORKBOOK, combined)

    added_2024_rows = sum(1 for r in combined if str(r.get("date", "")).startswith("2024-"))
    total_rows = len(combined)
    skipped_dates_total = sum(skipped.values())

    blank_2024_keys = [
        r
        for r in combined
        if str(r.get("date", "")).startswith("2024-")
        and (not r.get("stadium") or not r.get("away_team") or not r.get("home_team"))
    ]

    print(f"output_workbook: {OUTPUT_WORKBOOK}")
    print(f"base_rows_2025_2026: {base_count}")
    print(f"added_2024_rows: {added_2024_rows}")
    print(f"total_combined_rows: {total_rows}")
    print(f"included_2024_dates: {len(included_dates)}")
    print(f"skipped_2024_dates: {skipped_dates_total}")
    if included_dates:
        ordered_dates = sorted(included_dates)
        print(f"included_2024_date_first: {ordered_dates[0]}")
        print(f"included_2024_date_last: {ordered_dates[-1]}")
        print("included_2024_dates_csv:")
        print(",".join(ordered_dates))
    if skipped:
        print("skipped_summary:")
        for k, v in skipped.most_common():
            print(f"  {k}: {v}")
    print(f"blank_2024_stadium_away_home: {len(blank_2024_keys)}")


if __name__ == "__main__":
    main()
