from __future__ import annotations

import datetime as dt
import re
import time
from collections import Counter, defaultdict

import openpyxl
import requests
from bs4 import BeautifulSoup


INPUT_WORKBOOK = r"C:\Users\student\Documents\golfModel\park_factors_2026_season_to_date_ALL_FIXED.xlsx"
OUTPUT_WORKBOOK = r"C:\Users\student\Documents\golfModel\park_factors_2025_2026_ALL_FIXED.xlsx"

HEADERS = [
    "date",
    "stadium",
    "away_team",
    "home_team",
    "teams_playing",
    "hr_factor",
    "2b_3b_factor",
    "1b_factor",
    "runs_factor",
    "wind_receptive",
    "wind_hour1_mph",
    "wind_hour2_mph",
    "wind_hour3_mph",
    "air_hour1_temp",
    "air_hour2_temp",
    "air_hour3_temp",
    "humidity",
    "pressure",
]

UA = {
    "User-Agent": "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)"
}


def clean_value(value: str) -> str:
    return value.replace("\xa0", " ").strip()


def normalize_num(value: str) -> str:
    s = clean_value(value)
    if not s:
        return ""
    s = s.replace("%", "").replace("°", "").replace(",", "")
    if s in {"-", "—"}:
        return ""
    return s


def parse_game_blob(blob: str) -> tuple[str, str, str] | None:
    text = clean_value(blob)
    # Examples:
    # "Steinbrenner Field 7:05 PIT @ TB"
    # "Coors Field 8:40 NYM @ COL"
    m = re.search(r"^(.*?)\s+\d{1,2}:\d{2}\s+([A-Z]{2,3})\s*@\s*([A-Z]{2,3})$", text)
    if not m:
        return None
    stadium = clean_value(m.group(1))
    away = clean_value(m.group(2))
    home = clean_value(m.group(3))
    return stadium, away, home


def build_team_fallback(soup: BeautifulSoup) -> dict[str, tuple[str, str]]:
    park_to_teams: dict[str, set[str]] = defaultdict(set)
    tables = soup.find_all("table")
    if len(tables) < 2:
        return {}
    player_table = tables[1]
    rows = player_table.find_all("tr")
    for tr in rows[1:]:
        cells = [clean_value(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if len(cells) < 3:
            continue
        tm, park = cells[0], cells[2]
        if tm and park:
            park_to_teams[park].add(tm)
    out: dict[str, tuple[str, str]] = {}
    for park, teams in park_to_teams.items():
        if len(teams) >= 2:
            ordered = sorted(teams)
            out[park] = (ordered[0], ordered[1])
    return out


def parse_day(date_obj: dt.date) -> tuple[list[dict[str, str]], str | None]:
    date_str = date_obj.isoformat()
    url = f"https://www.ballparkpal.com/Park-Factors.php?date={date_str}"
    resp = requests.get(url, headers=UA, timeout=30)
    if resp.status_code != 200:
        return [], f"http_{resp.status_code}"

    soup = BeautifulSoup(resp.text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return [], "no_tables"

    game_table = tables[0]
    rows = game_table.find_all("tr")
    if len(rows) <= 1:
        return [], "no_game_rows"

    team_fallback = build_team_fallback(soup)
    parsed_rows: list[dict[str, str]] = []

    for tr in rows[1:]:
        cells = [clean_value(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if len(cells) < 17:
            continue
        if cells[0].lower() == "game":
            continue

        # Default structure in crawler HTML:
        # 0 game, 1 HR, 2 2B/3B, 3 1B, 4 Runs, 6 Receptive,
        # 8/9/10 wind mph, 11/12/13 temp, 15 humidity, 16 pressure
        game_blob = cells[0]
        game_parts = parse_game_blob(game_blob)
        if game_parts is None:
            # Fallback if game blob is missing and percentages start at col 1.
            # In that case no reliable game data in row; skip unless we can infer via park text.
            continue
        stadium, away, home = game_parts

        if not (stadium and away and home):
            # Team inference fallback from player table by stadium.
            fb = team_fallback.get(stadium)
            if fb:
                away, home = fb

        if not (stadium and away and home):
            continue

        row = {
            "date": date_str,
            "stadium": stadium,
            "away_team": away,
            "home_team": home,
            "teams_playing": f"{away} @ {home}",
            "hr_factor": normalize_num(cells[1]),
            "2b_3b_factor": normalize_num(cells[2]),
            "1b_factor": normalize_num(cells[3]),
            "runs_factor": normalize_num(cells[4]),
            "wind_receptive": clean_value(cells[6]),
            "wind_hour1_mph": normalize_num(cells[8]),
            "wind_hour2_mph": normalize_num(cells[9]),
            "wind_hour3_mph": normalize_num(cells[10]),
            "air_hour1_temp": normalize_num(cells[11]),
            "air_hour2_temp": normalize_num(cells[12]),
            "air_hour3_temp": normalize_num(cells[13]),
            "humidity": normalize_num(cells[15]),
            "pressure": normalize_num(cells[16]),
        }
        parsed_rows.append(row)

    if not parsed_rows:
        return [], "no_parsed_rows"
    return parsed_rows, None


def load_base_rows(path: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(HEADERS, start=1):
            v = ws.cell(r, c).value
            rec[h] = "" if v is None else str(v).strip()
        rows.append(rec)
    return rows


def dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str, str]] = set()
    out: list[dict[str, str]] = []
    for row in rows:
        key = (
            row.get("date", ""),
            row.get("stadium", ""),
            row.get("away_team", ""),
            row.get("home_team", ""),
        )
        if key in seen:
            continue
        if not all(key):
            continue
        seen.add(key)
        out.append(row)
    return out


def write_workbook(path: str, rows: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "park_factors"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])
    wb.save(path)


def main() -> None:
    base_rows = load_base_rows(INPUT_WORKBOOK)
    base_count = len(base_rows)

    start = dt.date(2025, 3, 1)
    end = dt.date(2025, 11, 15)

    all_2025_rows: list[dict[str, str]] = []
    included_dates: set[str] = set()
    skipped = Counter()

    d = start
    while d <= end:
        try:
            day_rows, reason = parse_day(d)
            if day_rows:
                all_2025_rows.extend(day_rows)
                included_dates.add(d.isoformat())
            else:
                skipped[reason or "unknown"] += 1
        except Exception:
            skipped["exception"] += 1
        d += dt.timedelta(days=1)
        time.sleep(0.25)

    combined = dedupe_rows(base_rows + all_2025_rows)
    write_workbook(OUTPUT_WORKBOOK, combined)

    added_2025_rows = sum(1 for r in combined if str(r.get("date", "")).startswith("2025-"))
    total_rows = len(combined)
    skipped_dates_total = sum(skipped.values())

    print(f"output_workbook: {OUTPUT_WORKBOOK}")
    print(f"base_rows_2026: {base_count}")
    print(f"added_2025_rows: {added_2025_rows}")
    print(f"total_combined_rows: {total_rows}")
    print(f"included_2025_dates: {len(included_dates)}")
    print(f"skipped_2025_dates: {skipped_dates_total}")
    if skipped:
        print("skipped_summary:")
        for k, v in skipped.most_common():
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
